import openpyxl

from ..engine import RedactionEngine
from .base import FileHandler


class XlsxHandler(FileHandler):
    def redact(
        self,
        engine: RedactionEngine,
        input_path: str,
        output_path: str,
        entity_types: list[str] | None = None,
        use_placeholders: bool = True,
    ) -> dict:
        wb = openpyxl.load_workbook(input_path)
        total_found = 0
        session_id = None

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None or not isinstance(cell.value, str):
                        continue
                    if not cell.value.strip():
                        continue
                    result = engine.redact(cell.value, entity_types=entity_types)
                    if result["entities_found"] > 0:
                        total_found += result["entities_found"]
                        if session_id is None:
                            session_id = result["session_id"]
                        else:
                            self._merge_session(engine, session_id, result["session_id"])
                        cell.value = result["redacted_text"]

        if session_id is None:
            session_id = engine.state.create_session()

        wb.save(output_path)
        return {"session_id": session_id, "entities_found": total_found}

    def unredact(
        self,
        input_path: str,
        output_path: str,
        mappings: dict[str, str],
    ) -> dict:
        wb = openpyxl.load_workbook(input_path)
        entities_restored = 0

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None or not isinstance(cell.value, str):
                        continue
                    new_value = cell.value
                    for placeholder, original in mappings.items():
                        if placeholder in new_value:
                            new_value = new_value.replace(placeholder, original)
                            entities_restored += 1
                    if new_value != cell.value:
                        cell.value = new_value

        wb.save(output_path)
        return {"entities_restored": entities_restored}

    @staticmethod
    def _merge_session(engine: RedactionEngine, target_id: str, source_id: str) -> None:
        source_mappings = engine.state.get_mappings(source_id)
        if source_mappings:
            for placeholder, original in source_mappings.items():
                engine.state.add_mapping(target_id, placeholder, original)
